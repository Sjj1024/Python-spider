# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html
import openpyxl


class JdphonePipeline(object):
    def open_spider(self, spider):  # 在爬虫开启的时候仅执行一次
        if spider.name == 'jdphone':
            self.filted_list = []  # 过滤出来的2019手机字典列表
            self.name_list = []  # 手机去重列表
            self.excel_dict = {"A": "品牌", "B": "型号", "C": "价格", "D": "上市时间", "E": "好评率", "F": "图片", "G": "详情链接",
                               "H": "机身长度", "I": "机身宽度", "J": "机身重量", "K": "屏幕尺寸"}
            # 没有过滤的2019手机信息表格
            self.f = openpyxl.Workbook()
            self.sheet1 = self.f.create_sheet("手机信息1")
            self.sheet1.append(self.excel_dict)
            # 2按照时间过滤后的信息表格
            self.five_list = []  # 保存价格在500以下的手机
            self.onefive_list = []  # 保存价格在1500以下的手机
            self.twofive_list = []  # 保存价格在2500以下的手机
            self.thrfive_list = []  # 保存价格在3500以下的手机
            self.forfive_list = []  # 保存价格在3500以上的手机
            self.f2 = openpyxl.Workbook()
            self.sheet2 = self.f2.create_sheet("手机信息2")
            self.sheet2.append(self.excel_dict)
            # 3按照手机品牌分类
            self.filter_brand = {}  # 创建一个字典，保存手机品牌对应的手机列表
            self.f3 = openpyxl.Workbook()
            self.sheet3 = self.f3.create_sheet("手机信息3")
            self.sheet3.append(self.excel_dict)

    def close_spider(self, spider):  # 在爬虫关闭的时候仅执行一次
        if spider.name == 'jdphone':
            # 保存2019没有过滤的手机列表到表格中
            self.f.save("2019手机总和详情1.xlsx")
            # 2.1然后按照时间对手机进行排序
            self.five_list = sorted(self.five_list, key=lambda x: int(x['month_time'][:1:]))
            self.onefive_list = sorted(self.onefive_list, key=lambda x: int(x['month_time'][:1:]))
            self.twofive_list = sorted(self.twofive_list, key=lambda x: int(x['month_time'][:1:]))
            self.thrfive_list = sorted(self.thrfive_list, key=lambda x: int(x['month_time'][:1:]))
            self.forfive_list = sorted(self.forfive_list, key=lambda x: int(x['month_time'][:1:]))
            self.all_list = [self.five_list, self.onefive_list, self.twofive_list, self.thrfive_list,
                             self.forfive_list]  # 用来存储排序后的手机列表
            # 遍历排序后的列表，然后存储到表格中
            for onelist in self.all_list:
                # 取出一个排序好的列表
                self.sheet2.append({"A": "                              "})
                self.sheet2.append({"A": "以下是价格在500到500的手机信息："})
                for item in onelist:
                    # 将单个手机字典对象添加到表格中
                    self.sheet2.append(
                        {"A": item['brand'], "B": item['name'], "C": item['price'],
                         "D": item['year_time'] + item['month_time'],
                         "E": item['goodrate'], "F": item['image'], "G": item['link'], "H": item["length"],
                         "I": item["width"],
                         "J": item["weight"], "K": item["inch"]})
            self.f2.save("2019手机价格排序2.xlsx")

            # 按照手机品牌进行分类
            # 按照时间先进行排序处理
            for name in self.filter_brand:
                self.filter_brand[name].sort(key=lambda x: int(x['month_time'][:1:]))

            # 然后保存到表格中
            for name in self.filter_brand:
                # 遍历手机品牌中的每一个商品
                # 取出一个排序好的列表
                self.sheet3.append({"A": "                              "})
                self.sheet3.append({"A": f"以下是{name}手机产品列表"})
                for item in self.filter_brand[name]:
                    # 将单个手机字典对象添加到表格中
                    self.sheet3.append(
                        {"A": item['brand'], "B": item['name'], "C": item['price'],
                         "D": item['year_time'] + item['month_time'],
                         "E": item['goodrate'], "F": item['image'], "G": item['link'], "H": item["length"],
                         "I": item["width"],
                         "J": item["weight"], "K": item["inch"]})
            self.f3.save("2019手机品牌分类3.xlsx")

    def process_item(self, item, spider):
        print("管道开始执行了00000000000000000000000000000000000000")
        # 过滤出2019年才发布的手机
        if item['year_time'] and item['year_time'][:4:] == "2019" and item['month_time'] != '以官网信息为准':
            # 再过滤出有名字的手机
            if item['name'] and item['name'] not in ['以官网信息为准', '·', '产品名称', '-', '--', '其他', '以官网数据为准', '官网信息为准',
                                                     '以官网为准']:
                # 再过滤出已经查询出来的手机
                if item['name'].upper().replace(" ", "") not in self.name_list:
                    # 再进行双向过滤出已经存在的手机
                    uppname = item['name'].upper().replace(" ", "")
                    if all(uppname not in element.upper().replace(" ", "") and element.upper().replace(" ",
                                                                                                       "") not in uppname
                           for element in self.name_list):
                        # 1.将过滤后的手机信息保存到表格中
                        self.sheet1.append(
                            {"A": item['brand'], "B": item['name'], "C": item['price'],
                             "D": item['year_time'] + item['month_time'],
                             "E": item['goodrate'], "F": item['image'], "G": item['link'], "H": item["length"],
                             "I": item["width"],
                             "J": item["weight"], "K": item["inch"]})
                        self.name_list.append(item['name'].upper().replace(" ", ""))
                        # 将过滤后的手机保存到字典列表中，
                        self.filted_list.append(item)
                        # 2.按照手机价格进行汇聚，
                        if float(item['price']) <= 500:
                            self.five_list.append(item)
                        elif 500 < float(item['price']) <= 1500:
                            self.onefive_list.append(item)
                        elif 1500 < float(item['price']) <= 2500:
                            self.twofive_list.append(item)
                        elif 2500 < float(item['price']) <= 3500:
                            self.thrfive_list.append(item)
                        elif 3500 < float(item['price']):
                            self.forfive_list.append(item)

                        # 判断手机品牌已存在，将手机品牌名字保存到列表字典中
                        if item["brand"] in self.filter_brand:
                            self.filter_brand[item["brand"]].append(item)
                        else:
                            self.filter_brand[item["brand"]] = [item]
        print("管道执行结束=============================================")
